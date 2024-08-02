import openpyxl

class FunctionEx:

    def __init__(self, driver):
        self.driver = driver
    def read_data(file, path, sheet_name, row_num, column_num):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_num, column=column_num).value
    def write_data(file, path, sheet_name, row_num, column_num, data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num, column=column_num).value = data
        workbook.save(path)

    def get_row_count(file, path, sheet_name):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    def get_column_count(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_column